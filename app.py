from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import os
import secrets
import hashlib
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# CORS-Konfiguration für React Frontend
CORS(app, 
     supports_credentials=True,
     origins=['http://localhost:5173', 'http://127.0.0.1:5173'],
     allow_headers=['Content-Type'],
     methods=['GET', 'POST', 'DELETE', 'OPTIONS'])

# Datenbank-Konfiguration
# Lokal: SQLite (keine Admin-Rechte nötig) | Production: PostgreSQL
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///dienstwuensche.db')
# Fix für Render: postgres:// -> postgresql+psycopg:// (für psycopg3)
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql+psycopg://', 1)
elif app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgresql://'):
    # Explizit psycopg3 Driver verwenden
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Datenbank-Modelle
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    email = db.Column(db.String(120), nullable=True)
    password = db.Column(db.String(64), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    force_password_change = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    shift_requests = db.relationship('ShiftRequest', backref='user', lazy=True, cascade='all, delete-orphan')
    first_submission_at = db.Column(db.DateTime, nullable=True)  # Zeitpunkt des ersten Speicherns

class ShiftRequest(db.Model):
    __tablename__ = 'shift_requests'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    shift_type = db.Column(db.String(20), nullable=False)
    remarks = db.Column(db.Text, default='')
    status = db.Column(db.String(20), default='PENDING')
    confirmed = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    shift_notes = db.relationship('ShiftNote', backref='shift_request', lazy=True, cascade='all, delete-orphan')

class ShiftNote(db.Model):
    __tablename__ = 'shift_notes'
    id = db.Column(db.Integer, primary_key=True)
    shift_request_id = db.Column(db.Integer, db.ForeignKey('shift_requests.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User')

class ShiftRequestSnapshot(db.Model):
    """Speichert ursprüngliche Dienstwünsche beim ersten Speichern"""
    __tablename__ = 'shift_request_snapshots'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    shift_type = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User')

class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', backref='messages')
    read_by = db.relationship('MessageRead', backref='message', lazy=True, cascade='all, delete-orphan')

class MessageRead(db.Model):
    __tablename__ = 'message_reads'
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.Integer, db.ForeignKey('messages.id'), nullable=False)
    admin_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    read_at = db.Column(db.DateTime, default=datetime.now)
    admin = db.relationship('User')

def hash_password(password):
    """Erstelle einen Hash des Passworts"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(stored_hash, password):
    """Überprüfe ob Passwort korrekt ist"""
    return stored_hash == hash_password(password)

def generate_temp_password(length=12):
    """Generiere ein temporäres Passwort"""
    import string
    import random
    chars = string.ascii_letters + string.digits + "!@#$%"
    return ''.join(random.choice(chars) for _ in range(length))

def get_current_user():
    """Hole den aktuell angemeldeten Benutzer"""
    user_name = session.get('user_name')
    if not user_name:
        return None
    return User.query.filter_by(name=user_name).first()

def is_admin():
    """Prüfe ob aktueller Benutzer Admin ist"""
    user = get_current_user()
    return user and user.is_admin

def require_login():
    """Prüfe ob Benutzer angemeldet ist"""
    if not get_current_user():
        return jsonify({'success': False, 'error': 'Nicht angemeldet'}), 401
    return None

def migrate_database():
    """Führt Datenbank-Migration durch"""
    from sqlalchemy import text, inspect
    
    def check_column_exists(table_name, column_name):
        inspector = inspect(db.engine)
        columns = [col['name'] for col in inspector.get_columns(table_name)]
        return column_name in columns
    
    def check_table_exists(table_name):
        inspector = inspect(db.engine)
        return table_name in inspector.get_table_names()
    
    # Erkenne Datenbanktyp
    is_postgres = 'postgresql' in str(db.engine.url)
    
    try:
        # Prüfe ob updated_at Spalte existiert
        if check_table_exists('shift_requests') and not check_column_exists('shift_requests', 'updated_at'):
            print("   Füge updated_at Spalte zu shift_requests hinzu...")
            with db.engine.connect() as conn:
                conn.execute(text("""
                    ALTER TABLE shift_requests 
                    ADD COLUMN updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                """))
                conn.commit()
            print("   ✓ updated_at Spalte hinzugefügt")
        
        # Prüfe ob shift_notes Tabelle existiert
        if not check_table_exists('shift_notes'):
            print("   Erstelle shift_notes Tabelle...")
            with db.engine.connect() as conn:
                if is_postgres:
                    # PostgreSQL Syntax
                    conn.execute(text("""
                        CREATE TABLE shift_notes (
                            id SERIAL PRIMARY KEY,
                            shift_request_id INTEGER NOT NULL,
                            user_id INTEGER NOT NULL,
                            content TEXT NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (shift_request_id) REFERENCES shift_requests(id) ON DELETE CASCADE,
                            FOREIGN KEY (user_id) REFERENCES users(id)
                        )
                    """))
                else:
                    # SQLite Syntax
                    conn.execute(text("""
                        CREATE TABLE shift_notes (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            shift_request_id INTEGER NOT NULL,
                            user_id INTEGER NOT NULL,
                            content TEXT NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (shift_request_id) REFERENCES shift_requests(id) ON DELETE CASCADE,
                            FOREIGN KEY (user_id) REFERENCES users(id)
                        )
                    """))
                conn.commit()
            print("   ✓ shift_notes Tabelle erstellt")
        
        # Prüfe ob first_submission_at Spalte in users existiert
        if check_table_exists('users') and not check_column_exists('users', 'first_submission_at'):
            print("   Füge first_submission_at Spalte zu users hinzu...")
            with db.engine.connect() as conn:
                conn.execute(text("""
                    ALTER TABLE users 
                    ADD COLUMN first_submission_at TIMESTAMP
                """))
                conn.commit()
            print("   ✓ first_submission_at Spalte hinzugefügt")
        
        # Prüfe ob shift_request_snapshots Tabelle existiert
        if not check_table_exists('shift_request_snapshots'):
            print("   Erstelle shift_request_snapshots Tabelle...")
            with db.engine.connect() as conn:
                if is_postgres:
                    conn.execute(text("""
                        CREATE TABLE shift_request_snapshots (
                            id SERIAL PRIMARY KEY,
                            user_id INTEGER NOT NULL,
                            date DATE NOT NULL,
                            shift_type VARCHAR(20) NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                        )
                    """))
                else:
                    conn.execute(text("""
                        CREATE TABLE shift_request_snapshots (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            user_id INTEGER NOT NULL,
                            date DATE NOT NULL,
                            shift_type VARCHAR(20) NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                        )
                    """))
                conn.commit()
            print("   ✓ shift_request_snapshots Tabelle erstellt")
    except Exception as e:
        print(f"   Warnung bei Migration: {e}")

def init_db():
    """Initialisiere Datenbank"""
    with app.app_context():
        db.create_all()
        # Führe Migration durch
        migrate_database()
        # Erstelle Initial-Admin falls keine Benutzer existieren
        if User.query.count() == 0:
            admin = User(
                name='Groß',
                password=hash_password('mettwurst'),
                is_admin=True
            )
            db.session.add(admin)
            db.session.commit()
            print("✓ Initial-Admin 'Groß' erstellt")

@app.route('/')
def index():
    """Hauptseite mit Formular"""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    # Admins zur Admin-Seite umleiten
    if user.is_admin:
        return redirect(url_for('admin_dashboard'))
    return render_template('shift_request_form_new.html', user_name=user.name)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login-Seite"""
    if request.method == 'POST':
        # Unterstütze sowohl JSON als auch Form-Data
        if request.is_json:
            data = request.json
        else:
            data = request.form
            
        name = data.get('name', '').strip()
        password = data.get('password', '')
        new_password = data.get('new_password', '')
        
        if not name:
            return jsonify({'success': False, 'error': 'Name ist erforderlich'}), 400
        
        if not password:
            return jsonify({'success': False, 'error': 'Passwort ist erforderlich'}), 400
        
        user = User.query.filter_by(name=name).first()
        
        if not user:
            return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 401
        
        # Passwort prüfen
        if not verify_password(user.password, password):
            return jsonify({'success': False, 'error': 'Falsches Passwort'}), 401
        
        # Prüfe ob Passwort geändert werden muss
        if user.force_password_change:
            if not new_password:
                return jsonify({
                    'success': False, 
                    'force_password_change': True,
                    'error': 'Sie müssen Ihr Passwort ändern'
                }), 403
            
            # Neues Passwort setzen
            user.password = hash_password(new_password)
            user.force_password_change = False
            db.session.commit()
        
        # Benutzer in Session speichern
        session['user_name'] = name
        
        # Für normale Formulare: Redirect
        if not request.is_json:
            return redirect(url_for('index'))
        
        return jsonify({'success': True, 'is_admin': user.is_admin})
    
    return render_template('login.html')

@app.route('/register', methods=['POST'])
def register():
    """Registrierung für neue Benutzer"""
    try:
        # Unterstütze sowohl JSON als auch Form-Data
        if request.is_json:
            data = request.json
        else:
            data = request.form
            
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        confirm_password = data.get('confirm_password', '')
        
        if not name:
            return jsonify({'success': False, 'error': 'Name ist erforderlich'}), 400
        
        if not password or len(password) < 6:
            return jsonify({'success': False, 'error': 'Passwort muss mindestens 6 Zeichen lang sein'}), 400
        
        if password != confirm_password:
            return jsonify({'success': False, 'error': 'Passwörter stimmen nicht überein'}), 400
        
        # Prüfe ob Benutzer bereits existiert
        if User.query.filter_by(name=name).first():
            return jsonify({'success': False, 'error': 'Dieser Name ist bereits vergeben'}), 400
        
        # Prüfe ob E-Mail bereits existiert (falls angegeben)
        if email and User.query.filter_by(email=email).first():
            return jsonify({'success': False, 'error': 'Diese E-Mail-Adresse ist bereits vergeben'}), 400
        
        # Erstelle neuen Benutzer
        new_user = User(
            name=name,
            email=email if email else None,
            password=hash_password(password),
            is_admin=False,
            force_password_change=False
        )
        db.session.add(new_user)
        db.session.commit()
        
        # Für normale Formulare: Redirect zum Login
        if not request.is_json:
            return redirect(url_for('login'))
        
        return jsonify({'success': True, 'message': 'Registrierung erfolgreich! Sie können sich jetzt anmelden.'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/logout')
def logout():
    """Logout"""
    session.pop('user_name', None)
    return redirect(url_for('login'))

@app.route('/admin')
def admin_dashboard():
    """Admin-Dashboard mit allen Dienstwünschen"""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    
    if not user.is_admin:
        return redirect(url_for('index'))
    
    # Monat-Filter aus Query-Parameter
    selected_month = request.args.get('month')
    selected_year = request.args.get('year')
    
    # Standard: Folgemonat
    if not selected_month or not selected_year:
        today = datetime.now()
        if today.month == 12:
            selected_month = 1
            selected_year = today.year + 1
        else:
            selected_month = today.month + 1
            selected_year = today.year
    else:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    
    # Lade alle Benutzer
    all_users = User.query.order_by(User.name).all()
    users_data = []
    for u in all_users:
        # Zähle shift_requests ohne sie komplett zu laden
        shift_count = ShiftRequest.query.filter_by(user_id=u.id).count()
        users_data.append({
            'id': u.id,
            'name': u.name,
            'is_admin': u.is_admin,
            'created_at': u.created_at.isoformat(),
            'shift_count': shift_count
        })
    
    # Lade Dienstwünsche für ausgewählten Monat
    all_requests = []
    
    # Ermittle welche User tatsächlich Änderungen haben
    users_with_modifications = set()
    for user in User.query.filter(User.first_submission_at.isnot(None)).all():
        # Hole Snapshots und aktuelle Shifts
        snapshots = ShiftRequestSnapshot.query.filter_by(user_id=user.id).all()
        current_shifts = ShiftRequest.query.filter_by(user_id=user.id).all()
        
        # Erstelle Sets für Vergleich
        snapshot_set = {(s.date.isoformat(), s.shift_type) for s in snapshots}
        current_set = {(s.date.isoformat(), s.shift_type) for s in current_shifts}
        
        # Prüfe ob es Unterschiede gibt
        if snapshot_set != current_set:
            users_with_modifications.add(user.id)
    
    for req in ShiftRequest.query.filter(
        db.extract('month', ShiftRequest.date) == selected_month,
        db.extract('year', ShiftRequest.date) == selected_year
    ).order_by(ShiftRequest.date).all():
        # Lade Notizen für diesen Dienst
        notes_data = []
        for note in req.shift_notes:
            notes_data.append({
                'id': note.id,
                'content': note.content,
                'user_name': note.user.name,
                'created_at': note.created_at.isoformat()
            })
        
        all_requests.append({
            'id': str(req.id),
            'user_id': req.user_id,
            'user_name': req.user.name,
            'date': req.date.isoformat(),
            'shiftType': req.shift_type,
            'remarks': req.remarks,
            'status': req.status,
            'confirmed': req.confirmed,
            'createdAt': req.created_at.isoformat(),
            'updatedAt': req.updated_at.isoformat() if req.updated_at else req.created_at.isoformat(),
            'first_submission_at': req.user.first_submission_at.isoformat() if req.user.first_submission_at else None,
            'has_modifications': req.user_id in users_with_modifications,
            'notes': notes_data
        })
    
    # Generiere Liste verfügbarer Monate (letzte 12 Monate + nächste 3)
    import calendar as cal
    available_months = []
    today = datetime.now()
    for i in range(-12, 4):
        month_date = datetime(today.year, today.month, 1) + timedelta(days=32 * i)
        month_date = month_date.replace(day=1)
        available_months.append({
            'year': month_date.year,
            'month': month_date.month,
            'name': cal.month_name[month_date.month],
            'display': f"{cal.month_name[month_date.month]} {month_date.year}"
        })
    
    return render_template('admin_dashboard_v3.html', 
                         requests=all_requests,
                         users=users_data,
                         user_name=user.name,
                         selected_month=selected_month,
                         selected_year=selected_year,
                         available_months=available_months)

# Admin API Endpoints
@app.route('/api/admin/users', methods=['GET'])
def get_all_users():
    """Hole alle Benutzer (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        users = User.query.order_by(User.name).all()
        users_data = []
        for u in users:
            users_data.append({
                'id': u.id,
                'name': u.name,
                'is_admin': u.is_admin,
                'created_at': u.created_at.isoformat(),
                'shift_count': len(u.shift_requests)
            })
        
        return jsonify({'success': True, 'data': users_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
def toggle_admin(user_id):
    """Admin-Rolle umschalten (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 404
        
        # Verhindere, dass der letzte Admin sich selbst degradiert
        if user.is_admin:
            admin_count = User.query.filter_by(is_admin=True).count()
            if admin_count <= 1:
                return jsonify({'success': False, 'error': 'Es muss mindestens ein Admin geben'}), 400
        
        user.is_admin = not user.is_admin
        db.session.commit()
        
        return jsonify({'success': True, 'is_admin': user.is_admin})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/users/<int:user_id>/reset-password', methods=['POST'])
def reset_password(user_id):
    """Passwort zurücksetzen (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 404
        
        # Generiere temporäres Passwort
        temp_password = generate_temp_password()
        user.password = hash_password(temp_password)
        user.force_password_change = True
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'temp_password': temp_password,
            'message': f'Neues Passwort für {user.name}: {temp_password}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/shift-requests/<int:request_id>/confirm', methods=['POST'])
def confirm_shift_request(request_id):
    """Dienstwunsch bestätigen (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        shift_request = ShiftRequest.query.get(request_id)
        if not shift_request:
            return jsonify({'success': False, 'error': 'Dienstwunsch nicht gefunden'}), 404
        
        shift_request.confirmed = not shift_request.confirmed
        db.session.commit()
        
        return jsonify({'success': True, 'confirmed': shift_request.confirmed})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/users/<int:user_id>/confirm-all-shifts', methods=['POST'])
def confirm_all_user_shifts(user_id):
    """Alle Dienstwünsche eines Benutzers bestätigen (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 404
        
        # Hole alle unbestätigten Dienstwünsche des Benutzers
        shifts = ShiftRequest.query.filter_by(user_id=user_id, confirmed=False).all()
        
        for shift in shifts:
            shift.confirmed = True
        
        db.session.commit()
        
        return jsonify({'success': True, 'confirmed_count': len(shifts)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-notes', methods=['POST'])
def create_shift_note():
    """Erstelle Notiz zu einem Dienst"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        user = get_current_user()
        data = request.json
        
        shift_id = data.get('shift_id')
        content = data.get('content', '').strip()
        
        if not shift_id or not content:
            return jsonify({'success': False, 'error': 'Dienst-ID und Inhalt erforderlich'}), 400
        
        # Prüfe ob Dienst existiert
        shift_request = ShiftRequest.query.get(shift_id)
        if not shift_request:
            return jsonify({'success': False, 'error': 'Dienst nicht gefunden'}), 404
        
        note = ShiftNote(shift_request_id=shift_id, user_id=user.id, content=content)
        db.session.add(note)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'note': {
                'id': note.id,
                'content': note.content,
                'user_name': user.name,
                'created_at': note.created_at.isoformat()
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-notes/<int:shift_id>', methods=['GET'])
def get_shift_notes(shift_id):
    """Hole alle Notizen zu einem Dienst"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        notes = ShiftNote.query.filter_by(shift_request_id=shift_id).order_by(ShiftNote.created_at).all()
        
        notes_data = []
        for note in notes:
            notes_data.append({
                'id': note.id,
                'content': note.content,
                'user_name': note.user.name,
                'created_at': note.created_at.isoformat()
            })
        
        return jsonify({'success': True, 'notes': notes_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/messages', methods=['GET', 'POST'])
def messages():
    """Nachrichten senden und abrufen"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    user = get_current_user()
    
    if request.method == 'POST':
        # Neue Nachricht senden
        data = request.json if request.is_json else request.form
        content = data.get('content', '').strip()
        
        if not content:
            return jsonify({'success': False, 'error': 'Nachricht darf nicht leer sein'}), 400
        
        try:
            message = Message(user_id=user.id, content=content)
            db.session.add(message)
            db.session.commit()
            
            return jsonify({'success': True, 'message_id': message.id})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    else:
        # Nachrichten abrufen (nur für Admins)
        if not is_admin():
            return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
        
        messages_data = []
        all_messages = Message.query.order_by(Message.created_at.desc()).all()
        
        for msg in all_messages:
            # Prüfe welche Admins diese Nachricht gelesen haben
            read_by_admins = []
            for read in msg.read_by:
                read_by_admins.append({
                    'admin_id': read.admin_id,
                    'admin_name': read.admin.name,
                    'read_at': read.read_at.isoformat()
                })
            
            # Prüfe ob aktueller Admin gelesen hat
            has_read = any(read.admin_id == user.id for read in msg.read_by)
            
            messages_data.append({
                'id': msg.id,
                'user_id': msg.user_id,
                'user_name': msg.user.name,
                'content': msg.content,
                'created_at': msg.created_at.isoformat(),
                'has_read': has_read,
                'read_by': read_by_admins
            })
        
        return jsonify({'success': True, 'messages': messages_data})

@app.route('/api/messages/<int:message_id>/read', methods=['POST'])
def mark_message_read(message_id):
    """Nachricht als gelesen markieren"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    user = get_current_user()
    
    try:
        message = Message.query.get(message_id)
        if not message:
            return jsonify({'success': False, 'error': 'Nachricht nicht gefunden'}), 404
        
        # Prüfe ob bereits gelesen
        existing = MessageRead.query.filter_by(message_id=message_id, admin_id=user.id).first()
        if existing:
            return jsonify({'success': True, 'already_read': True})
        
        # Als gelesen markieren
        read_entry = MessageRead(message_id=message_id, admin_id=user.id)
        db.session.add(read_entry)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/export')
def export_shift_requests():
    """Exportiere alle Dienstwünsche als JSON"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        # Gruppiere Dienstwünsche nach Nutzern
        users = User.query.order_by(User.name).all()
        export_data = []
        
        for user in users:
            user_requests = []
            for req in user.shift_requests:
                user_requests.append({
                    'date': req.date.strftime('%d.%m.%Y'),
                    'shift_type': req.shift_type,
                    'remarks': req.remarks,
                    'confirmed': req.confirmed
                })
            
            if user_requests:  # Nur Benutzer mit Dienstwünschen
                export_data.append({
                    'user_name': user.name,
                    'shift_requests': user_requests
                })
        
        return jsonify({'success': True, 'data': export_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/export/excel')
def export_excel():
    """Exportiere Dienstwünsche als Excel-Datei"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        # Erstelle Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dienstwünsche"
        
        # Styles
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        headers = ['Mitarbeiter', 'Datum', 'Wochentag', 'Schichtart', 'Bemerkungen', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Daten
        row = 2
        all_requests = ShiftRequest.query.order_by(ShiftRequest.user_id, ShiftRequest.date).all()
        
        for req in all_requests:
            ws.cell(row=row, column=1, value=req.user.name).border = border
            ws.cell(row=row, column=2, value=req.date.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row, column=3, value=req.date.strftime('%A')).border = border
            ws.cell(row=row, column=4, value=req.shift_type).border = border
            ws.cell(row=row, column=5, value=req.remarks or '').border = border
            status_cell = ws.cell(row=row, column=6, value='Bestätigt' if req.confirmed else 'Offen')
            status_cell.border = border
            if req.confirmed:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            row += 1
        
        # Spaltenbreiten
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        
        # Speichern in BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Dienstwünsche_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/export/pdf')
def export_pdf():
    """Exportiere Dienstwünsche als PDF-Datei mit Kalender-Layout"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        import calendar
        
        buffer = BytesIO()
        
        # Berechne Folgemonat
        today = datetime.now()
        if today.month == 12:
            next_month = 1
            next_year = today.year + 1
        else:
            next_month = today.month + 1
            next_year = today.year
        
        # Hole alle Dienstwünsche für Folgemonat
        all_shifts = ShiftRequest.query.filter(
            db.extract('month', ShiftRequest.date) == next_month,
            db.extract('year', ShiftRequest.date) == next_year
        ).all()
        
        # Gruppiere nach Benutzer und Datum
        user_shifts = {}
        for shift in all_shifts:
            if shift.user.name not in user_shifts:
                user_shifts[shift.user.name] = {}
            day = shift.date.day
            user_shifts[shift.user.name][day] = {
                'type': shift.shift_type,
                'confirmed': shift.confirmed
            }
        
        # Sortiere Benutzer
        sorted_users = sorted(user_shifts.keys())
        
        # Erstelle PDF
        c = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)
        
        # Titel
        month_name = calendar.month_name[next_month]
        c.setFont("Helvetica-Bold", 16)
        c.setFillColorRGB(0.75, 0, 0)  # DRK Rot
        c.drawCentredString(width/2, height - 40, f"Dienstplan {month_name} {next_year}")
        
        # Untertitel
        c.setFont("Helvetica", 10)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(width/2, height - 60, f"Erstellt am {today.strftime('%d.%m.%Y %H:%M')}")
        
        # Kalender-Tabelle
        days_in_month = calendar.monthrange(next_year, next_month)[1]
        
        # Spaltenbreiten
        name_col_width = 100
        day_col_width = (width - 80 - name_col_width) / days_in_month
        
        # Startposition
        x_start = 40
        y_start = height - 100
        row_height = 25
        
        # Header mit Tagesnummern
        c.setFont("Helvetica-Bold", 8)
        for day in range(1, days_in_month + 1):
            x = x_start + name_col_width + (day - 1) * day_col_width
            c.drawCentredString(x + day_col_width/2, y_start, str(day))
        
        # Wochentage unter den Tagesnummern
        c.setFont("Helvetica", 6)
        for day in range(1, days_in_month + 1):
            date_obj = datetime(next_year, next_month, day)
            weekday = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'][date_obj.weekday()]
            x = x_start + name_col_width + (day - 1) * day_col_width
            c.drawCentredString(x + day_col_width/2, y_start - 10, weekday)
        
        # Mitarbeiter-Zeilen
        y_pos = y_start - 30
        c.setFont("Helvetica", 9)
        
        for user_name in sorted_users:
            # Name
            c.drawString(x_start, y_pos, user_name)
            
            # Schichten für jeden Tag
            for day in range(1, days_in_month + 1):
                x = x_start + name_col_width + (day - 1) * day_col_width
                
                if day in user_shifts[user_name]:
                    shift_info = user_shifts[user_name][day]
                    shift_type = shift_info['type']
                    
                    # Hintergrundfarbe je nach Schichttyp
                    if shift_type == 'T':
                        c.setFillColorRGB(0.99, 0.95, 0.78)  # Gelb
                    elif shift_type == 'T10':
                        c.setFillColorRGB(0.86, 0.92, 0.99)  # Blau
                    elif shift_type == 'N10':
                        c.setFillColorRGB(0.88, 0.91, 1.0)  # Indigo
                    else:
                        c.setFillColorRGB(0.9, 0.9, 0.9)  # Grau
                    
                    c.rect(x, y_pos - 5, day_col_width, row_height, fill=1, stroke=0)
                    
                    # Text
                    c.setFillColorRGB(0, 0, 0)
                    c.setFont("Helvetica-Bold", 7)
                    c.drawCentredString(x + day_col_width/2, y_pos + 5, shift_type)
                    
                    # Bestätigt-Marker
                    if shift_info['confirmed']:
                        c.setFont("Helvetica", 6)
                        c.setFillColorRGB(0, 0.5, 0)
                        c.drawCentredString(x + day_col_width/2, y_pos - 2, '✓')
                
                # Rahmen
                c.setStrokeColorRGB(0.7, 0.7, 0.7)
                c.setLineWidth(0.5)
                c.rect(x, y_pos - 5, day_col_width, row_height, fill=0, stroke=1)
            
            y_pos -= row_height
            
            # Neue Seite wenn nötig
            if y_pos < 100:
                c.showPage()
                y_pos = height - 100
                
                # Header wiederholen
                c.setFont("Helvetica-Bold", 8)
                for day in range(1, days_in_month + 1):
                    x = x_start + name_col_width + (day - 1) * day_col_width
                    c.drawCentredString(x + day_col_width/2, y_pos + 30, str(day))
                
                c.setFont("Helvetica", 6)
                for day in range(1, days_in_month + 1):
                    date_obj = datetime(next_year, next_month, day)
                    weekday = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'][date_obj.weekday()]
                    x = x_start + name_col_width + (day - 1) * day_col_width
                    c.drawCentredString(x + day_col_width/2, y_pos + 20, weekday)
        
        # Legende
        y_legend = 60
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x_start, y_legend, "Legende:")
        
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(0.99, 0.95, 0.78)
        c.rect(x_start + 60, y_legend - 5, 20, 12, fill=1, stroke=1)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(x_start + 85, y_legend, "T = Tagdienst")
        
        c.setFillColorRGB(0.86, 0.92, 0.99)
        c.rect(x_start + 170, y_legend - 5, 20, 12, fill=1, stroke=1)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(x_start + 195, y_legend, "T10 = Tagdienst 10h")
        
        c.setFillColorRGB(0.88, 0.91, 1.0)
        c.rect(x_start + 300, y_legend - 5, 20, 12, fill=1, stroke=1)
        c.setFillColorRGB(0, 0, 0)
        c.drawString(x_start + 325, y_legend, "N10 = Nachtdienst 10h")
        
        c.setFillColorRGB(0, 0.5, 0)
        c.drawString(x_start + 450, y_legend, "✓ = Bestätigt")
        
        # Fußzeile
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawCentredString(width/2, 30, "🏥 DRK Köln - Erste-Hilfe-Station Flughafen")
        
        c.save()
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Dienstplan_{month_name}_{next_year}.pdf'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-requests', methods=['GET'])
def get_shift_requests():
    """Hole alle Dienstwünsche des Benutzers für den Folgemonat"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        user = get_current_user()
        # Berechne Folgemonat
        today = datetime.now()
        if today.month == 12:
            next_month = 1
            next_year = today.year + 1
        else:
            next_month = today.month + 1
            next_year = today.year
        
        # Filtere nach Benutzer und Folgemonat
        requests = ShiftRequest.query.filter_by(user_id=user.id).filter(
            db.extract('month', ShiftRequest.date) == next_month,
            db.extract('year', ShiftRequest.date) == next_year
        ).order_by(ShiftRequest.date).all()
        
        filtered = []
        for req in requests:
            # Lade Notizen für diesen Dienst
            notes_data = []
            for note in req.shift_notes:
                notes_data.append({
                    'id': note.id,
                    'content': note.content,
                    'user_name': note.user.name,
                    'created_at': note.created_at.isoformat()
                })
            
            filtered.append({
                'id': str(req.id),
                'user_name': user.name,
                'date': req.date.isoformat(),
                'shiftType': req.shift_type,
                'remarks': req.remarks,
                'status': req.status,
                'confirmed': req.confirmed,
                'createdAt': req.created_at.isoformat(),
                'updatedAt': req.updated_at.isoformat() if req.updated_at else req.created_at.isoformat(),
                'notes': notes_data
            })
        
        return jsonify({'success': True, 'data': filtered})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-requests', methods=['POST'])
def create_shift_request():
    """Erstelle einen neuen Dienstwunsch"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        user = get_current_user()
        data = request.json
        
        # Validierung
        if not data.get('date'):
            return jsonify({'success': False, 'error': 'Datum ist erforderlich'}), 400
        
        if not data.get('shiftType'):
            return jsonify({'success': False, 'error': 'Schichtart ist erforderlich'}), 400
        
        # Prüfe ob Datum in der Vergangenheit liegt
        request_date = datetime.fromisoformat(data['date'])
        if request_date.date() < datetime.now().date():
            return jsonify({
                'success': False, 
                'error': 'Das Datum darf nicht in der Vergangenheit liegen'
            }), 400
        
        # Prüfe ob bereits ein Wunsch für diesen Tag und Benutzer existiert
        existing = ShiftRequest.query.filter_by(
            user_id=user.id,
            date=request_date.date()
        ).first()
        
        if existing:
            return jsonify({
                'success': False,
                'error': 'Sie haben bereits einen Wunsch für dieses Datum eingereicht'
            }), 400
        
        # Erstelle neuen Wunsch
        new_request = ShiftRequest(
            user_id=user.id,
            date=request_date.date(),
            shift_type=data['shiftType'],
            remarks=data.get('remarks', ''),
            status='PENDING'
        )
        
        db.session.add(new_request)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'data': {
                'id': str(new_request.id),
                'user_name': user.name,
                'date': new_request.date.isoformat(),
                'shiftType': new_request.shift_type,
                'remarks': new_request.remarks,
                'status': new_request.status,
                'confirmed': new_request.confirmed,
                'createdAt': new_request.created_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/users/<int:user_id>/snapshots', methods=['GET'])
def get_user_snapshots(user_id):
    """Hole ursprüngliche Dienstwünsche eines Benutzers (nur Admin)"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    if not is_admin():
        return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
    
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 404
        
        snapshots = ShiftRequestSnapshot.query.filter_by(user_id=user_id).order_by(ShiftRequestSnapshot.date).all()
        current_shifts = ShiftRequest.query.filter_by(user_id=user_id).order_by(ShiftRequest.date).all()
        
        return jsonify({
            'success': True,
            'user_name': user.name,
            'first_submission_at': user.first_submission_at.isoformat() if user.first_submission_at else None,
            'snapshots': [{'date': s.date.isoformat(), 'shift_type': s.shift_type} for s in snapshots],
            'current': [{
                'date': s.date.isoformat(), 
                'shift_type': s.shift_type, 
                'confirmed': s.confirmed,
                'created_at': s.created_at.isoformat(),
                'updated_at': s.updated_at.isoformat() if s.updated_at else s.created_at.isoformat()
            } for s in current_shifts]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-requests/batch', methods=['POST'])
def save_shifts_batch():
    """Speichere mehrere Dienstwünsche gleichzeitig mit Änderungsverfolgung"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        user = get_current_user()
        data = request.json
        shifts = data.get('shifts', {})  # Dict: {date: shiftType}
        
        # Prüfe ob User bereits einmal gespeichert hat
        is_first_submission = user.first_submission_at is None
        has_changes = False
        
        if not is_first_submission:
            # Prüfe ob es Änderungen gibt (neue Dienste oder geänderte Dienste)
            existing_shifts = {sr.date.isoformat(): sr.shift_type for sr in ShiftRequest.query.filter_by(user_id=user.id).all()}
            
            for date_str, shift_type in shifts.items():
                if date_str not in existing_shifts or existing_shifts[date_str] != shift_type:
                    has_changes = True
                    break
            
            # Prüfe auch ob Dienste entfernt wurden
            for date_str in existing_shifts:
                if date_str not in shifts and not ShiftRequest.query.filter_by(user_id=user.id, date=datetime.fromisoformat(date_str).date(), confirmed=True).first():
                    has_changes = True
                    break
        
        # Lösche alle nicht-bestätigten Dienstwünsche des Users
        ShiftRequest.query.filter_by(user_id=user.id, confirmed=False).delete()
        
        # Erstelle neue Dienstwünsche
        new_shifts = []
        for date_str, shift_type in shifts.items():
            # Überspringe wenn bereits bestätigt
            if ShiftRequest.query.filter_by(user_id=user.id, date=datetime.fromisoformat(date_str).date(), confirmed=True).first():
                continue
            
            new_shift = ShiftRequest(
                user_id=user.id,
                date=datetime.fromisoformat(date_str).date(),
                shift_type=shift_type,
                status='PENDING'
            )
            db.session.add(new_shift)
            new_shifts.append(new_shift)
        
        # Bei erster Einreichung: Setze Zeitstempel und erstelle Snapshots
        if is_first_submission:
            user.first_submission_at = datetime.now()
            
            # Erstelle Snapshots der ursprünglichen Dienste
            for date_str, shift_type in shifts.items():
                snapshot = ShiftRequestSnapshot(
                    user_id=user.id,
                    date=datetime.fromisoformat(date_str).date(),
                    shift_type=shift_type
                )
                db.session.add(snapshot)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'is_modification': not is_first_submission and has_changes,
            'shift_count': len(new_shifts)
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/shift-requests/<request_id>', methods=['DELETE'])
def delete_shift_request(request_id):
    """Lösche einen Dienstwunsch"""
    auth_error = require_login()
    if auth_error:
        return auth_error
    
    try:
        user = get_current_user()
        
        # Finde den Wunsch
        shift_request = ShiftRequest.query.get(int(request_id))
        
        if not shift_request:
            return jsonify({'success': False, 'error': 'Wunsch nicht gefunden'}), 404
        
        # Prüfe ob der Wunsch dem aktuellen Benutzer gehört
        if shift_request.user_id != user.id:
            return jsonify({'success': False, 'error': 'Nicht autorisiert'}), 403
        
        # Lösche den Wunsch
        db.session.delete(shift_request)
        db.session.commit()
        
        return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    print("\n✅ Dienstwunsch-Anwendung startet...")
    print("🌐 Öffne im Browser: http://localhost:5000")
    print("⛔ Zum Beenden: STRG + C\n")
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
